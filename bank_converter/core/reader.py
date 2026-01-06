"""
File readers for bank export data.

=== LEARNING: This module demonstrates ===
- Context managers (with statement) - like RAII in C++
- Exception handling (try/except) - like try/catch in C++
- Optional type hints - like std::optional in C++
- Iterating over files and collections
- Working with the csv module
- Lazy imports for optional dependencies
"""

# =============================================================================
# IMPORTS
# =============================================================================
import csv                              # Built-in CSV parsing module
from pathlib import Path                # Object-oriented path handling
from typing import List, Optional       # Type hints for complex types

# === LEARNING: Relative imports ===
# The dot (.) means "current package" (same as this file's directory)
# [C++] Similar to: #include "models.h" (local include)
#       vs. #include <vector> (system include)
#
# from .models    = from bank_converter/core/models.py
# from ..config   = from bank_converter/config/ (parent's sibling)
from .models import BankTransaction


class BankFileReader:
    """
    Reads bank export files (CSV or XLSX).

    === LEARNING: Encapsulation in Python ===
    Python uses "we're all adults here" philosophy.
    There are no private/protected/public keywords!

    [C++] No equivalent to:
          private:
              std::vector<std::string> expected_columns;
          public:
              std::vector<Transaction> read(const std::string& path);

    Instead, Python uses NAMING CONVENTIONS:
    - public_method()     - normal name, meant to be used
    - _private_method()   - underscore prefix = "internal use only"
    - __mangled_method()  - double underscore = name mangling (rarely used)

    The underscore is just a hint - nothing stops you from calling _methods!
    """

    # =========================================================================
    # CLASS VARIABLE (static member)
    # =========================================================================
    # === LEARNING: Class Variables ===
    # [C++] Like: static const std::vector<std::string> EXPECTED_COLUMNS;
    # [JS] Like: static EXPECTED_COLUMNS = [...]
    #
    # This is shared by ALL instances of BankFileReader.
    # Accessed as: BankFileReader.EXPECTED_COLUMNS or self.EXPECTED_COLUMNS
    EXPECTED_COLUMNS = [
        "Дата операции", "Дата платежа", "Номер карты", "Статус",
        "Сумма операции", "Валюта операции", "Сумма платежа",
        "Валюта платежа", "Кэшбэк", "Категория", "MCC", "Описание"
    ]

    # =========================================================================
    # PUBLIC METHOD
    # =========================================================================
    def read(self, file_path: str) -> List[BankTransaction]:
        """
        Read file and return list of transactions.

        === LEARNING: Docstrings ===
        Triple-quoted strings right after function definition are "docstrings".
        They become part of the function: read.__doc__

        [C++] Like Doxygen comments, but accessible at runtime:
              /// @param file_path Path to bank export file
              /// @return List of transactions
              /// @throws ValueError if format not supported

        Args:
            file_path: Path to bank export file (CSV or XLSX).

        Returns:
            List of BankTransaction objects.

        Raises:
            ValueError: If file format is not supported.
        """
        # === LEARNING: Path object ===
        # [C++] Similar to std::filesystem::path (C++17)
        # Provides OS-independent path manipulation
        path = Path(file_path)

        # === LEARNING: Method chaining and string methods ===
        # path.suffix returns file extension like ".csv" or ".XLSX"
        # .lower() converts to lowercase for case-insensitive comparison
        # [C++] Like: std::filesystem::path(file_path).extension()
        #       then boost::algorithm::to_lower(ext)
        if path.suffix.lower() == '.xlsx':
            return self._read_xlsx(path)
        elif path.suffix.lower() == '.csv':
            return self._read_csv(path)
        else:
            # === LEARNING: Raising Exceptions ===
            # [C++] Like: throw std::invalid_argument("message");
            # [JS] Like: throw new Error("message");
            #
            # Python has many built-in exception types:
            # - ValueError: wrong value for an operation
            # - TypeError: wrong type
            # - KeyError: missing dictionary key
            # - FileNotFoundError: file doesn't exist
            # - ImportError: module not found
            raise ValueError(f"Unsupported file format: {path.suffix}")

    # =========================================================================
    # PRIVATE METHODS (by convention - note the underscore prefix)
    # =========================================================================
    def _read_csv(self, path: Path) -> List[BankTransaction]:
        """
        Read semicolon-separated CSV with UTF-8 encoding.

        Tries multiple encodings because bank exports may vary.
        """
        transactions = []  # Empty list to collect results

        # === LEARNING: for loop over a sequence ===
        # [C++] Like: for (const auto& encoding : encodings) { ... }
        # Python's for loop ALWAYS iterates over something (no C-style for)
        for encoding in ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251']:

            # === LEARNING: try/except (Exception Handling) ===
            # [C++] Like: try { ... } catch (const UnicodeError& e) { ... }
            #
            # Key differences from C++:
            # - "except" instead of "catch"
            # - "raise" instead of "throw"
            # - No need to catch by reference (Python handles object lifetime)
            # - Can catch multiple exception types in one except block
            try:
                # === LEARNING: Context Manager (with statement) - CRITICAL! ===
                # This is Python's RAII equivalent!
                #
                # [C++] This is like:
                #       {
                #           std::ifstream f(path);  // Opens file
                #           // ... use f ...
                #       }  // Destructor closes file automatically (RAII)
                #
                # The 'with' statement guarantees:
                # 1. __enter__() is called (opens file, returns file object)
                # 2. Block executes
                # 3. __exit__() is called (closes file) - EVEN IF EXCEPTION OCCURS!
                #
                # Without 'with', you'd need:
                #     f = open(path, 'r')
                #     try:
                #         content = f.read()
                #     finally:
                #         f.close()  # Must remember to close!
                with open(path, 'r', encoding=encoding) as f:

                    # === LEARNING: csv.DictReader ===
                    # Reads CSV where first row is headers, returns dicts.
                    # [C++] No standard equivalent - you'd parse manually or use library
                    #
                    # Each row becomes: {"column1": "value1", "column2": "value2"}
                    # Instead of: ["value1", "value2"]
                    reader = csv.DictReader(f, delimiter=';', quotechar='"')

                    # === LEARNING: Iterating over reader ===
                    # csv.DictReader is an ITERATOR - reads one row at a time.
                    # [C++] Like using input iterator: while (reader >> row) { ... }
                    # This is memory-efficient: entire file isn't loaded at once!
                    for row in reader:
                        tx = self._parse_row(row)

                        # === LEARNING: Truthy/Falsy values ===
                        # 'if tx:' checks if tx is "truthy"
                        # In Python, these are "falsy": None, 0, "", [], {}, False
                        # Everything else is "truthy"
                        # [C++] Like: if (tx != nullptr && tx->isValid())
                        if tx:
                            transactions.append(tx)

                    # Return on success (exit the encoding loop)
                    return transactions

            except UnicodeDecodeError:
                # Wrong encoding, try next one
                # [C++] Like: catch (const std::runtime_error& e) { continue; }
                continue

            except Exception as e:
                # === LEARNING: 'as' binds exception to variable ===
                # [C++] Like: catch (const std::exception& e) { ... }
                #
                # 'e' contains the exception object with message, traceback, etc.
                # e.args[0] usually has the message
                # str(e) converts to string representation
                if encoding == 'windows-1251':
                    # Last encoding failed - re-raise the exception
                    # [C++] Like: throw; (re-throw current exception)
                    raise e
                continue

        # If we get here, all encodings failed
        raise ValueError("Could not read file with any supported encoding")

    def _read_xlsx(self, path: Path) -> List[BankTransaction]:
        """Read XLSX file using openpyxl."""

        # === LEARNING: Lazy Import ===
        # Import inside function instead of at top of file.
        # [C++] No equivalent - all #includes must be at file level
        #
        # Why do this?
        # 1. openpyxl is optional - don't fail if not installed but not needed
        # 2. Faster startup - module only loaded when actually needed
        # 3. Avoids circular import issues
        try:
            import openpyxl
        except ImportError:
            # === LEARNING: Custom exception messages ===
            # Provide helpful error messages with solution!
            raise ImportError(
                "openpyxl is required for XLSX support. "
                "Install it with: pip install openpyxl"
            )

        # === LEARNING: Library-specific patterns ===
        # openpyxl.load_workbook returns a Workbook object
        # read_only=True: faster for reading, uses less memory
        # data_only=True: get cell values, not formulas
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active  # Get the active (first) worksheet
        transactions = []

        # === LEARNING: Converting iterator to list ===
        # sheet.iter_rows() returns a generator (lazy iterator)
        # list(...) forces it to load all rows into memory
        # [C++] Like: std::vector<Row> rows(iter.begin(), iter.end());
        #
        # values_only=True: get cell values, not Cell objects
        rows = list(sheet.iter_rows(values_only=True))

        # Early return if empty
        if not rows:
            return transactions

        # === LEARNING: List comprehension for transformation ===
        # [C++] Like: std::transform with lambda
        #       std::vector<std::string> headers;
        #       std::transform(rows[0].begin(), rows[0].end(),
        #                      std::back_inserter(headers),
        #                      [](auto cell) { return cell ? std::to_string(cell) : ""; });
        #
        # The ternary: str(cell) if cell else ""
        # [C++] Like: cell ? std::to_string(cell) : ""
        headers = [str(cell) if cell else "" for cell in rows[0]]

        # === LEARNING: Slicing ===
        # rows[1:] = all rows starting from index 1 (skip header)
        # [C++] No direct equivalent. You'd use:
        #       for (size_t i = 1; i < rows.size(); i++)
        #       or iterators: for (auto it = rows.begin() + 1; it != rows.end(); ++it)
        for row in rows[1:]:
            # Skip empty rows
            # all() returns True if ALL items are truthy
            # [C++] Like: std::all_of(row.begin(), row.end(),
            #                         [](auto c) { return c == nullptr; })
            if not row or all(cell is None for cell in row):
                continue

            # Build dictionary from headers and row values
            row_dict = {}

            # === LEARNING: enumerate() ===
            # Returns (index, value) pairs when iterating
            # [C++] Like using a counter:
            #       for (size_t i = 0; i < headers.size(); i++) {
            #           const auto& header = headers[i];
            #       }
            for i, header in enumerate(headers):
                # Bounds checking with ternary
                # [C++] Like: i < row.size() ? row[i] : nullptr
                value = row[i] if i < len(row) else None
                row_dict[header] = str(value) if value is not None else ""

            tx = self._parse_row(row_dict)
            if tx:
                transactions.append(tx)

        # Explicitly close workbook (good practice, though not strictly required)
        workbook.close()
        return transactions

    def _parse_row(self, row: dict) -> Optional[BankTransaction]:
        """
        Parse row dict into BankTransaction.

        === LEARNING: Optional type hint ===
        Optional[BankTransaction] means "BankTransaction or None"
        [C++] Like: std::optional<BankTransaction>

        This tells callers: "Check the return value - it might be None!"
        """
        # Skip empty rows using dict.get() for safe access
        # .get(key) returns None if key doesn't exist (no KeyError)
        # [C++] Like: row.count("Дата операции") && !row["Дата операции"].empty()
        if not row or not row.get("Дата операции"):
            return None

        # === LEARNING: Named arguments in constructor ===
        # Python allows passing arguments by name in any order!
        # [C++] No equivalent - must pass args in order, or use builder pattern:
        #       BankTransaction::builder()
        #           .operationDate(...)
        #           .paymentDate(...)
        #           .build();
        return BankTransaction(
            operation_date=self._clean_str(row.get("Дата операции", "")),
            payment_date=self._clean_str(row.get("Дата платежа", "")),
            card_number=self._clean_str(row.get("Номер карты", "")),
            status=self._clean_str(row.get("Статус", "")),
            amount=self._clean_str(row.get("Сумма операции", "")),
            currency=self._clean_str(row.get("Валюта операции", "")),
            category=self._clean_str(row.get("Категория", "")),
            mcc=self._clean_str(row.get("MCC", "")),
            description=self._clean_str(row.get("Описание", ""))
        )

    def _clean_str(self, value: str) -> str:
        """
        Clean string value - remove quotes, strip whitespace.

        === LEARNING: Method chaining ===
        [C++] Like chaining methods in fluent interfaces:
              value.strip().trimQuotes()

        Python strings are immutable, so each method returns a NEW string.
        """
        if not value:
            return ""

        # str(value) ensures we have a string (defensive programming)
        # .strip() removes whitespace from both ends
        # .strip('"') removes quote characters from both ends
        # [C++] Like: boost::algorithm::trim(value); followed by remove quotes
        return str(value).strip().strip('"')


# =============================================================================
# MODULE-LEVEL FUNCTIONS (not in a class)
# =============================================================================
# === LEARNING: Functions outside classes ===
# Python allows "free functions" like C++, not everything needs a class!
# [C++] Just like regular functions in a namespace:
#       namespace reader {
#           float parse_amount(const std::string& amount_str);
#       }
# [JS] Similar to: export function parseAmount(amountStr) { ... }

def parse_amount(amount_str: str) -> float:
    """
    Parse Russian-formatted amount string to float.

    === LEARNING: Docstring format ===
    This is Google-style docstring format.
    Other common formats: NumPy style, Sphinx style.

    Args:
        amount_str: Amount string like "-4363,00" or "142000.00"

    Returns:
        Float value of the amount.
    """
    # Falsy check - empty string, None, etc.
    if not amount_str:
        return 0.0

    # === LEARNING: Method chaining on strings ===
    # Each method returns a new string (immutable)
    # [C++] Like:
    #       std::string cleaned = amount_str;
    #       boost::algorithm::trim(cleaned);
    #       boost::algorithm::replace_all(cleaned, ",", ".");
    #       boost::algorithm::erase_all(cleaned, " ");
    cleaned = str(amount_str).strip('"').replace(',', '.').replace(' ', '')

    # '\xa0' is non-breaking space (common in formatted numbers)
    # [C++] Would be: cleaned.erase(std::remove(cleaned.begin(), cleaned.end(), '\xa0'), cleaned.end());
    cleaned = cleaned.replace('\xa0', '')

    # === LEARNING: Exception handling for conversion ===
    # [C++] Like using std::stof with try/catch:
    #       try {
    #           return std::stof(cleaned);
    #       } catch (const std::exception&) {
    #           return 0.0f;
    #       }
    try:
        return float(cleaned)
    except ValueError:
        # float() raises ValueError if string isn't a valid number
        return 0.0


def parse_date(date_str: str) -> str:
    """
    Convert various date formats to 'DD.MM.YYYY'.

    Supported inputs:
    - 'DD.MM.YYYY'
    - 'DD.MM.YYYY HH:MM:SS'
    - 'YYYY-MM-DD'

    Returns:
        Date string in DD.MM.YYYY format.
    """
    if not date_str:
        return ""

    # Defensive: ensure string type and clean
    date_str = str(date_str).strip().strip('"')

    # Remove time if present
    if ' ' in date_str:
        date_str = date_str.split(' ')[0]

    # Handle ISO format: YYYY-MM-DD
    if '-' in date_str:
        # [C++] split + index mapping
        # '2025-12-28' -> ['2025', '12', '28']
        parts = date_str.split('-')
        if len(parts) == 3:
            yyyy, mm, dd = parts
            return f"{dd}.{mm}.{yyyy}"

    # Already in DD.MM.YYYY (or unknown but left untouched)
    return date_str